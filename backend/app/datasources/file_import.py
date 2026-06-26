"""
File-import DataSource -- the always-available, first-class offline path.

Two input shapes are supported behind one interface:

  1. The spec workbook itself ("Copy of USA INV CHK.xlsx").  We read the same
     sheets a buyer pastes into today (INV, SALES, PRICE LIST, INC INV, plus
     the category sheets SEA/AIR/BLOOM/DOMESTIC/CLOTHING for category & source).

  2. Standalone Odoo exports the buyer downloads today:
       * an INV export  (Internal Reference, Cost, Free To Use Qty, On Hand)
       * a SALES export (product, Qty Ordered)  -- optionally month-by-month
     plus the PRICE LIST and INC INV files.

Either way it emits the common PullResult contract.  The workbook's SALES
sheet only carries a flat annual->monthly figure, so when no month-by-month
sales file is supplied the forecaster will (correctly) fall back to the flat
baseline and mark those SKUs low-confidence.  Supply a long-format monthly
CSV (global_sku,year,month,units[,is_stockout]) to unlock real forecasting.
"""
from __future__ import annotations

import csv
import os
import warnings
from typing import Dict, List, Optional

import openpyxl

from .base import PullResult
from ..config import (SOURCE_DOMESTIC, SOURCE_IMPORT_SEA)

warnings.filterwarnings("ignore", module="openpyxl")


def _s(v) -> str:
    return "" if v is None else str(v).strip()


def _f(v) -> Optional[float]:
    try:
        if v is None or v == "" or v == "#N/A":
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def load_from_workbook(path: str,
                       monthly_sales_csv: Optional[str] = None) -> PullResult:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    res = PullResult(source="file_import")

    # ---------------- PRICE LIST: mapping, COGS, retail, HSN -------------
    price: Dict[str, dict] = {}          # by global sku
    cogs_by_skucode: Dict[str, float] = {}  # PRICE LIST col R(SKU) -> col S(COGS)
    if "PRICE LIST" in wb.sheetnames:
        ws = wb["PRICE LIST"]
        for r in range(2, ws.max_row + 1):
            g = _s(ws.cell(r, 5).value)          # E US Global SKU
            if g and g != "#N/A":
                price[g] = {
                    "hsn_code": _s(ws.cell(r, 2).value),     # B HSN
                    "us_sku": _s(ws.cell(r, 6).value),       # F US SKU
                    "cogs": _f(ws.cell(r, 10).value),        # J LC USA
                    "retail_price": _f(ws.cell(r, 12).value),# L US RP
                    "name": _s(ws.cell(r, 4).value),         # D SKU Name
                }
            sk = _s(ws.cell(r, 18).value)         # R SKU
            cv = _f(ws.cell(r, 19).value)         # S COGS
            if sk and cv is not None:
                cogs_by_skucode[sk] = cv

    # ---------------- INV: on-hand, cost, us sku, name -------------------
    inv: Dict[str, dict] = {}
    if "INV" in wb.sheetnames:
        ws = wb["INV"]
        for r in range(3, ws.max_row + 1):
            ref = _s(ws.cell(r, 4).value)          # D Internal Reference
            if not ref or ref.startswith("???") or ref == "Internal Reference":
                continue
            inv[ref] = {
                "name": _s(ws.cell(r, 2).value),               # B prod name
                "cost": _f(ws.cell(r, 5).value) or 0.0,        # E Average Cost
                "free_to_use": _f(ws.cell(r, 7).value) or 0.0, # G Free To Use
                "on_hand": _f(ws.cell(r, 12).value) or 0.0,    # L Qty On Hand
                "us_sku": _s(ws.cell(r, 9).value),             # I Name(=US SKU)
                "sales_price": _f(ws.cell(r, 11).value) or 0.0,# K Sales Price
            }

    # ---------------- SALES: flat monthly + annual -----------------------
    sales: Dict[str, dict] = {}
    if "SALES" in wb.sheetnames:
        ws = wb["SALES"]
        for r in range(2, ws.max_row + 1):
            code = _s(ws.cell(r, 1).value)         # A product code
            mon = _f(ws.cell(r, 3).value)          # C monthly (annual/12)
            ann = _f(ws.cell(r, 5).value)          # E Qty Ordered annual
            if code and code not in ("#VALUE!",) and (mon or ann):
                sales[code] = {"monthly": mon or (ann / 12 if ann else 0.0),
                               "annual": ann or (mon * 12 if mon else 0.0)}

    # ---------------- categories & source from the category sheets -------
    category: Dict[str, str] = {}
    source: Dict[str, str] = {}
    expiry: Dict[str, bool] = {}
    vendor: Dict[str, str] = {}
    moq: Dict[str, float] = {}
    case_override: Dict[str, int] = {}
    target_override: Dict[str, float] = {}
    candidates: set = set()

    def add_cat(sheet, sku_col, cat_col, src=SOURCE_IMPORT_SEA,
                cat_const=None, expiry_flag=False, target_col=None,
                case_const=None):
        if sheet not in wb.sheetnames:
            return
        ws = wb[sheet]
        for r in range(2, ws.max_row + 1):
            g = _s(ws.cell(r, sku_col).value)
            if not g or g == "#N/A":
                continue
            candidates.add(g)
            category[g] = cat_const or _s(ws.cell(r, cat_col).value) or category.get(g, "")
            source.setdefault(g, src)
            if expiry_flag:
                expiry[g] = True
            if case_const:
                case_override[g] = case_const
            if target_col:
                t = _f(ws.cell(r, target_col).value)
                if t is not None:
                    target_override[g] = t

    # SEA is the primary candidate universe (col B sku, C cat, J target MOH)
    add_cat("SEA", 2, 3, target_col=10)
    add_cat("AIR", 3, 4, target_col=10)
    add_cat("BLOOM", 2, 3, cat_const="BLOOM", expiry_flag=True, case_const=32)
    add_cat("CLOTHING", 2, None, cat_const="CLOTHING", case_const=8)
    # DOMESTIC: col B sku, C vendor, G MOQ
    if "DOMESTIC" in wb.sheetnames:
        ws = wb["DOMESTIC"]
        for r in range(2, ws.max_row + 1):
            g = _s(ws.cell(r, 2).value)
            if not g or g == "#N/A":
                continue
            candidates.add(g)
            category[g] = "DOMESTIC"
            source[g] = SOURCE_DOMESTIC
            vendor[g] = _s(ws.cell(r, 3).value)
            mq = _f(ws.cell(r, 7).value)
            if mq is not None:
                moq[g] = mq

    # ---------------- optional month-by-month sales (long CSV) -----------
    monthly_series: Dict[str, List[dict]] = {}
    if monthly_sales_csv and os.path.exists(monthly_sales_csv):
        with open(monthly_sales_csv) as fh:
            for row in csv.DictReader(fh):
                g = _s(row.get("global_sku") or row.get("sku"))
                if not g:
                    continue
                monthly_series.setdefault(g, []).append({
                    "year": int(row["year"]),
                    "month": int(row["month"]),
                    "units": float(row.get("units") or row.get("qty") or 0),
                    "is_stockout": _s(row.get("is_stockout")).lower()
                    in ("1", "true", "yes"),
                })

    # ---------------- INC INV: in-transit by arrival month ---------------
    in_transit: Dict[tuple, dict] = {}
    if "INC INV" in wb.sheetnames:
        ws = wb["INC INV"]
        hdr = [(_s(ws.cell(1, c).value)).upper() for c in range(1, ws.max_column + 1)]
        # find repeating blocks: a SKU header with QTY and MONTH +2/+3 over
        c = 1
        while c <= ws.max_column:
            h = hdr[c - 1] if c - 1 < len(hdr) else ""
            if h == "SKU":
                lbl_col = c - 1            # SHIPMENT is one col before SKU
                qty_col = c + 2            # SKU, NAME, QTY
                mon_col = c + 3            # MONTH
                for r in range(2, ws.max_row + 1):
                    g = _s(ws.cell(r, c).value)
                    q = _f(ws.cell(r, qty_col).value)
                    m = _f(ws.cell(r, mon_col).value)
                    if not g or q is None or m is None:
                        continue
                    mi = int(round(m))
                    if mi < 1 or mi > 6:
                        continue
                    label = _s(ws.cell(r, lbl_col).value) if lbl_col >= 1 else ""
                    key = (g, mi)
                    rec = in_transit.setdefault(
                        key, {"global_sku": g, "expected_arrival_month": mi,
                              "quantity": 0.0, "shipment_label": label})
                    rec["quantity"] += q
                    if label and not rec["shipment_label"]:
                        rec["shipment_label"] = label
                c += 4
            else:
                c += 1

    # ---------------- assemble products + snapshots ----------------------
    # Candidate universe = the curated category sheets only (SEA/AIR/BLOOM/
    # DOMESTIC/CLOTHING).  This mirrors the workbook's order universe; we do
    # NOT pull in every SKU that merely has sales, which would add thousands
    # of uncategorised, non-ordered rows.
    for g in sorted(candidates):
        iv = inv.get(g, {})
        pr = price.get(g, {})
        sl = sales.get(g, {})
        cogs = pr.get("cogs")
        if cogs is None:
            cogs = cogs_by_skucode.get(iv.get("us_sku", ""), iv.get("cost", 0.0))
        retail = pr.get("retail_price") or iv.get("sales_price", 0.0)
        name = iv.get("name") or pr.get("name") or ""
        us_sku = iv.get("us_sku") or pr.get("us_sku", "")

        res.products.append({
            "global_sku": g,
            "us_sku": us_sku,
            "odoo_internal_ref": g,
            "name": name,
            "category": category.get(g, ""),
            "case_size": case_override.get(g, 1),
            "hsn_code": pr.get("hsn_code", ""),
            "cost": cogs or 0.0,
            "retail_price": retail or 0.0,
            "compliance_flag": "",
            "source": source.get(g, SOURCE_IMPORT_SEA),
            "expiry_tracked": expiry.get(g, False),
            "moq": moq.get(g),
            "target_moh_override": target_override.get(g),
            "vendor": vendor.get(g, ""),
        })

        on_hand = iv.get("free_to_use", iv.get("on_hand", 0.0))
        avg = sl.get("monthly", 0.0)
        res.snapshots.append({
            "global_sku": g,
            "on_hand": on_hand,
            "useable_on_hand": on_hand,
            "avg_monthly_sales": avg,
            "monthly_sales_series": monthly_series.get(g, []),
            "source": "file_import",
        })

    res.in_transit = list(in_transit.values())
    if not monthly_series:
        res.warnings.append(
            "No month-by-month sales file supplied: forecasting falls back to "
            "the flat annual/12 baseline (low confidence). Provide a long-format "
            "monthly CSV or use the live Odoo pull for seasonal forecasting.")
    return res


# --------------------------------------------------------------------------
# Standalone export path (INV.csv + SALES.csv + optional pricelist / inc)
# --------------------------------------------------------------------------
def load_from_exports(inv_path: str, sales_path: str,
                      price_path: Optional[str] = None,
                      inc_path: Optional[str] = None,
                      monthly_sales_csv: Optional[str] = None) -> PullResult:
    """Parse the discrete Odoo export files (CSV/XLSX) the buyer downloads
    today.  Column detection is tolerant of header naming."""
    res = PullResult(source="file_import")
    inv_rows = _read_table(inv_path)
    sales_rows = _read_table(sales_path)
    price_rows = _read_table(price_path) if price_path else []
    inc_rows = _read_table(inc_path) if inc_path else []

    def col(row, *names):
        for n in names:
            for k in row:
                if k and k.strip().lower() == n.lower():
                    return row[k]
        return None

    sales = {}
    for row in sales_rows:
        code = _parse_code(_s(col(row, "Product", "product", "code") or ""))
        qty = _f(col(row, "Qty Ordered", "qty ordered", "Total", "units"))
        if code and qty is not None:
            sales[code] = sales.get(code, 0.0) + qty

    price = {}
    for row in price_rows:
        g = _s(col(row, "US Global SKU", "global_sku", "SKU"))
        if g:
            price[g] = {"cogs": _f(col(row, "LC USA", "COGS", "cost")),
                        "retail_price": _f(col(row, "US RP", "retail")),
                        "hsn_code": _s(col(row, "HSN Code", "hsn")),
                        "us_sku": _s(col(row, "US SKU", "us_sku"))}

    seen = set()
    for row in inv_rows:
        ref = _s(col(row, "Internal Reference", "internal reference", "ref"))
        if not ref or ref in seen:
            continue
        seen.add(ref)
        pr = price.get(ref, {})
        on_hand = _f(col(row, "Free To Use Quantity", "free to use quantity",
                         "Quantity On Hand")) or 0.0
        res.products.append({
            "global_sku": ref, "us_sku": pr.get("us_sku", ""),
            "odoo_internal_ref": ref,
            "name": _s(col(row, "Display Name", "Name", "name")),
            "category": _s(col(row, "Product Category", "category")),
            "case_size": 1, "hsn_code": pr.get("hsn_code", ""),
            "cost": pr.get("cogs") or _f(col(row, "Cost")) or 0.0,
            "retail_price": pr.get("retail_price") or _f(col(row, "Sales Price")) or 0.0,
            "compliance_flag": "", "source": SOURCE_IMPORT_SEA,
            "expiry_tracked": False, "moq": None,
            "target_moh_override": None, "vendor": "",
        })
        res.snapshots.append({
            "global_sku": ref, "on_hand": on_hand, "useable_on_hand": on_hand,
            "avg_monthly_sales": (sales.get(ref, 0.0) / 12.0),
            "monthly_sales_series": [], "source": "file_import",
        })

    for row in inc_rows:
        g = _s(col(row, "SKU", "global_sku"))
        q = _f(col(row, "QTY", "qty", "quantity"))
        m = _f(col(row, "MONTH", "month"))
        if g and q and m:
            res.in_transit.append({
                "global_sku": g, "quantity": q,
                "expected_arrival_month": int(round(m)),
                "shipment_label": _s(col(row, "SHIPMENT", "label"))})
    return res


def _parse_code(display: str) -> str:
    """Pull CODE out of Odoo '[CODE] Name' display format."""
    s = display.strip()
    if s.startswith("[") and "]" in s:
        return s[1:s.index("]")].strip()
    return s


def _read_table(path: str) -> List[dict]:
    if not path or not os.path.exists(path):
        return []
    if path.lower().endswith((".xlsx", ".xls")):
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        hdr = [(_s(h) or f"c{i}") for i, h in enumerate(rows[0])]
        return [dict(zip(hdr, r)) for r in rows[1:]]
    with open(path, newline="") as fh:
        return list(csv.DictReader(fh))
