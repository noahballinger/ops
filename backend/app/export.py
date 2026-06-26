"""
Order export -- CSV and XLSX, columns mirroring the workbook's ORDER LIST so
output can be handed straight to the India team and the customs broker.

Columns (ORDER LIST parity + provenance):
  US SKU, NAME, GLOBAL SKU, CATEGORY, SEA QTY, AIR QTY, UNIT WEIGHT (G),
  UNIT COST (COGS), RETAIL, MARGIN, HSN, AIR SHIPPING COST, PROFIT LOST BY AIR,
  TARGET MOH, CASE, COMPLIANCE FLAG, SOURCE
The exported quantities are the *finalised* (override-aware) values.
"""
from __future__ import annotations

import csv
import io
from typing import List, Dict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

EXPORT_COLUMNS = [
    ("us_sku", "US SKU"),
    ("name", "NAME"),
    ("global_sku", "GLOBAL SKU"),
    ("category", "CATEGORY"),
    ("final_sea_qty", "SEA QTY"),
    ("final_air_qty", "AIR QTY"),
    ("unit_weight", "UNIT WEIGHT (G)"),
    ("unit_cost", "UNIT COST (COGS)"),
    ("retail_price", "RETAIL"),
    ("margin", "MARGIN"),
    ("hsn_code", "HSN"),
    ("air_shipping_cost", "AIR SHIPPING COST"),
    ("profit_lost_by_air", "PROFIT LOST BY AIR"),
    ("target_moh", "TARGET MOH"),
    ("case_size", "CASE"),
    ("compliance_flag", "COMPLIANCE FLAG"),
    ("source", "SOURCE"),
]


def rows_to_csv(rows: List[Dict]) -> str:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([h for _, h in EXPORT_COLUMNS])
    for r in rows:
        w.writerow([_fmt(r.get(k, "")) for k, _ in EXPORT_COLUMNS])
    return buf.getvalue()


def rows_to_xlsx(rows: List[Dict], order_name: str = "Order") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (order_name or "Order")[:31]
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF")
    for c, (_, h) in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(1, c, h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
    flag_fill = PatternFill("solid", fgColor="FFF2CC")
    for ri, r in enumerate(rows, start=2):
        for c, (k, _) in enumerate(EXPORT_COLUMNS, start=1):
            ws.cell(ri, c, _xl_val(r.get(k, "")))
        if r.get("compliance_flag"):
            for c in range(1, len(EXPORT_COLUMNS) + 1):
                ws.cell(ri, c).fill = flag_fill
    # widths
    widths = [12, 42, 16, 14, 9, 9, 12, 13, 9, 10, 12, 14, 16, 11, 7, 26, 16]
    for i, wdt in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = wdt
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fmt(v):
    return "" if v is None else v


def _xl_val(v):
    if v is None:
        return ""
    return v
