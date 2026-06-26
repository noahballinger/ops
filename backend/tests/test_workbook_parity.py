"""
Acceptance test: drive the pure engine with the workbook SEA sheet's OWN
inputs (MON SALES, MOH, target MOH, the six incoming-MOH columns) and assert
our SEA QTY / AIR QTY match the workbook's computed SEA QTY / AIR QTY within
rounding, across every fully-numeric row.

Set ISHA_WORKBOOK to the workbook path to run (skipped otherwise so CI without
the file still passes).
"""
import os
import sys

import pytest

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from app.config import EngineConfig
from app.engine import ProductInput, SkuSnapshot, suggest_one

WB = os.environ.get("ISHA_WORKBOOK")
pytestmark = pytest.mark.skipif(not (WB and os.path.exists(WB)),
                                reason="set ISHA_WORKBOOK to the spec workbook")


def _load_sea_rows():
    import openpyxl, warnings
    warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(WB, data_only=True)
    ws = wb["SEA"]
    rows = []
    # cols: F6 mon, H8 moh, J10 target, Q17 seaqty, S19 airqty, V-AA 22..27
    for r in range(2, ws.max_row + 1):
        def g(i):
            return ws.cell(r, i).value
        mon, moh, target = g(6), g(8), g(10)
        seaqty, airqty = g(17), g(19)
        inc = [g(i) for i in range(22, 28)]
        vals = [mon, moh, target, seaqty, airqty] + inc
        if all(isinstance(v, (int, float)) for v in vals) and mon and mon > 0:
            rows.append((g(1), mon, moh, target, seaqty, airqty, inc))
    return rows


def test_engine_matches_workbook_sea_air():
    cfg = EngineConfig()
    rows = _load_sea_rows()
    assert len(rows) >= 20, f"expected many comparable rows, got {len(rows)}"
    mismatches = []
    for (name, mon, moh, target, seaqty, airqty, inc) in rows:
        on_hand = moh * mon
        inc_units = [m * mon for m in inc]
        p = ProductInput(global_sku="X", category="X", cost=1, retail_price=2,
                         target_moh_override=target)
        snap = SkuSnapshot(product=p, on_hand=on_hand, avg_monthly_sales=mon,
                           incoming_units_by_month=inc_units, forecast=None)
        s = suggest_one(snap, cfg)
        if abs(s.suggested_sea_qty - seaqty) > 1.0 or \
           abs(s.suggested_air_qty - airqty) > 1.0:
            mismatches.append((name, s.suggested_sea_qty, seaqty,
                               s.suggested_air_qty, airqty))
    assert not mismatches, f"{len(mismatches)}/{len(rows)} rows differ: {mismatches[:5]}"
    print(f"\nPARITY OK: {len(rows)} SEA rows reproduced within rounding.")
